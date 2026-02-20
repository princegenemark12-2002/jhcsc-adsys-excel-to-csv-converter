import io
import os
import sys
from typing import Optional

import pandas as pd


def excel_to_csv(xlsx_path: str, csv_path: str) -> None:
    df = pd.read_excel(xlsx_path, header=None)
    df.to_csv(csv_path, index=False, header=False)


def format_english_header(xlsx_path: str, output_path: Optional[str] = None) -> None:
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment
        from openpyxl.utils import column_index_from_string
    except ImportError:
        print("openpyxl is required for formatting. Install with: pip install openpyxl")
        raise SystemExit(1)

    wb = load_workbook(xlsx_path)
    ws = wb.active

    col_end = column_index_from_string("BU")
    row_index: Optional[int] = None
    col_start: Optional[int] = None

    for row in range(1, 6):
        for cell in ws[row]:
            value = str(cell.value).strip().lower() if cell.value is not None else ""
            if value == "english":
                row_index = cell.row
                col_start = cell.column
                break
        if row_index is not None:
            break

    if row_index is None or col_start is None:
        print("Could not find an 'English' header to format.")
        return

    ws.merge_cells(
        start_row=row_index,
        start_column=col_start,
        end_row=row_index,
        end_column=col_end,
    )
    cell = ws.cell(row=row_index, column=col_start)
    cell.value = "English"
    cell.alignment = Alignment(horizontal="center", vertical="center")

    if output_path is None:
        if xlsx_path.lower().endswith(".xlsx"):
            output_path = xlsx_path[:-5] + "_merged.xlsx"
        else:
            output_path = xlsx_path + "_merged.xlsx"

    wb.save(output_path)
    print(f"Saved formatted workbook to {output_path}")


def run_web() -> None:
    try:
        from flask import Flask, request, send_file, render_template_string
    except ImportError:
        print("Flask is required for web mode. Install with: pip install flask")
        raise SystemExit(1)

    app = Flask(__name__)

    template = """
    <!doctype html>
    <html lang="en">
    <head>
        <meta charset="utf-8">
        <title>Excel to CSV Converter</title>
        <meta name="viewport" content="width=device-width, initial-scale=1">
        <style>
            * { box-sizing: border-box; margin: 0; padding: 0; }
            body {
                font-family: system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
                min-height: 100vh;
                display: flex;
                align-items: center;
                justify-content: center;
                background: linear-gradient(135deg, #0f172a, #1d4ed8);
                color: #0f172a;
            }
            .wrapper { width: 100%; max-width: 480px; padding: 24px; }
            .card {
                background: #f9fafb;
                border-radius: 16px;
                padding: 24px 20px;
                box-shadow: 0 18px 40px rgba(15,23,42,0.45);
            }
            .header { display: flex; align-items: center; gap: 10px; margin-bottom: 16px; }
            .logo-circle {
                width: 38px; height: 38px; border-radius: 999px;
                background: radial-gradient(circle at 30% 20%, #e0f2fe, #1d4ed8);
                display: flex; align-items: center; justify-content: center;
                color: #eff6ff; font-weight: 700; font-size: 18px;
            }
            h1 { font-size: 19px; color: #0f172a; }
            .subtitle { font-size: 12px; color: #6b7280; margin-top: 2px; }
            .file-label { font-size: 13px; margin-bottom: 6px; color: #4b5563; }
            .file-input-wrapper {
                position: relative;
                border-radius: 12px;
                border: 1px dashed #9ca3af;
                padding: 14px 12px;
                background: #f3f4f6;
                display: flex;
                align-items: center;
                justify-content: space-between;
                gap: 10px;
            }
            .file-input-wrapper span {
                font-size: 12px;
                color: #6b7280;
                overflow: hidden;
                text-overflow: ellipsis;
                white-space: nowrap;
            }
            .file-input-wrapper button {
                border: none;
                border-radius: 999px;
                background: #1d4ed8;
                color: #eff6ff;
                padding: 7px 12px;
                font-size: 12px;
                font-weight: 500;
                cursor: pointer;
            }
            .file-input-wrapper input[type=file] {
                position: absolute;
                inset: 0;
                opacity: 0;
                cursor: pointer;
            }
            .submit-row { margin-top: 16px; display: flex; justify-content: flex-end; }
            .submit-row button {
                border: none;
                border-radius: 999px;
                background: #16a34a;
                color: #ecfdf3;
                padding: 9px 18px;
                font-size: 13px;
                font-weight: 600;
                cursor: pointer;
                box-shadow: 0 10px 20px rgba(22,163,74,0.4);
            }
            .error {
                margin-top: 10px;
                font-size: 12px;
                color: #b91c1c;
                background: #fee2e2;
                border-radius: 8px;
                padding: 6px 8px;
            }
            .helper {
                margin-top: 10px;
                font-size: 11px;
                color: #9ca3af;
                text-align: right;
            }
        </style>
    </head>
    <body>
        <div class="wrapper">
            <div class="card">
                <div class="header">
                    <div class="logo-circle">XC</div>
                    <div>
                        <h1>Excel to CSV</h1>
                        <div class="subtitle">Upload Excel, download clean CSV in one click.</div>
                    </div>
                </div>
                <form method="post" enctype="multipart/form-data">
                    <label class="file-label">Excel file (.xlsx or .xls)</label>
                    <div class="file-input-wrapper">
                        <span id="file-name">No file chosen</span>
                        <button type="button">Browse</button>
                        <input id="file-input" type="file" name="file" accept=".xlsx,.xls" required>
                    </div>
                    {% if error %}
                    <div class="error">{{ error }}</div>
                    {% endif %}
                    <div class="submit-row">
                        <button type="submit">Convert and download</button>
                    </div>
                    <div class="helper">Your file is processed locally on this server.</div>
                </form>
            </div>
        </div>
        <script>
            const input = document.getElementById("file-input");
            const label = document.getElementById("file-name");
            if (input && label) {
                input.addEventListener("change", function () {
                    if (this.files && this.files.length > 0) {
                        label.textContent = this.files[0].name;
                    } else {
                        label.textContent = "No file chosen";
                    }
                });
            }
        </script>
    </body>
    </html>
    """

    @app.route("/", methods=["GET", "POST"])
    def index():
        if request.method == "GET":
            return render_template_string(template)

        uploaded_file = request.files.get("file")
        if not uploaded_file or uploaded_file.filename == "":
            return render_template_string(template, error="Please choose an Excel file.")

        df = pd.read_excel(uploaded_file, header=None)
        output = io.StringIO()
        df.to_csv(output, index=False, header=False)
        output.seek(0)

        filename = uploaded_file.filename.rsplit(".", 1)[0] + ".csv"

        return send_file(
            io.BytesIO(output.getvalue().encode("utf-8")),
            as_attachment=True,
            download_name=filename,
            mimetype="text/csv",
        )

    port = int(os.environ.get("PORT", "5000"))
    app.run(host="0.0.0.0", port=port)


def main() -> None:
    if len(sys.argv) == 2 and sys.argv[1].lower() == "web":
        run_web()
        return

    if len(sys.argv) == 3 and sys.argv[1].lower() == "format":
        format_english_header(sys.argv[2])
        return

    if len(sys.argv) != 3:
        print("Usage: python excel_to_csv.py input.xlsx output.csv")
        print("Or   : python excel_to_csv.py web")
        print("Or   : python excel_to_csv.py format input.xlsx")
        raise SystemExit(1)

    input_file = sys.argv[1]
    output_file = sys.argv[2]
    excel_to_csv(input_file, output_file)
    print("Conversion complete.")


if __name__ == "__main__":
    main()

